"""AIE Backend — cXML parsing, scenario transformation, Excel generation.
Run: uvicorn aie_server:app --port 8100 --reload
"""
from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import List, Optional
import json
import os
from pathlib import Path
from lxml import etree
from copy import deepcopy
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from aie_retrieval import retrieve_pos_sync

app = FastAPI(title="AIE Backend", version="1.0.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

DATA_DIR = Path(__file__).parent / "data"
DATA_DIR.mkdir(exist_ok=True)


@app.get("/health")
def health():
    return {"status": "ok", "service": "AIE Backend", "version": "1.0.0"}


class GenerateScenariosRequest(BaseModel):
    basePO: str
    scenarios: List[str]


class GenerateScenariosResponse(BaseModel):
    success: bool
    scenarios: List[dict]
    excelReady: bool
    excelPath: Optional[str] = None


@app.post("/api/aie/generate-scenarios", response_model=GenerateScenariosResponse)
def generate_scenarios(req: GenerateScenariosRequest):
    """Generate PO scenario variants from base PO cXML."""
    base_cxml_path = DATA_DIR / f"PO-{req.basePO}.cxml"

    if not base_cxml_path.exists():
        # Use sample if no file present
        base_cxml_path = DATA_DIR / "PO-7800459912.cxml"

    if not base_cxml_path.exists():
        return GenerateScenariosResponse(
            success=False, scenarios=[], excelReady=False,
        )

    base_xml = base_cxml_path.read_text(encoding="utf-8")
    results = []
    scenario_cxmls = {}

    for scenario_id in req.scenarios:
        transformed = apply_transformation(base_xml, scenario_id)
        scenario_name = SCENARIO_NAMES.get(scenario_id, scenario_id)
        output_path = DATA_DIR / f"scenario-{scenario_id}.cxml"
        output_path.write_text(transformed, encoding="utf-8")
        scenario_cxmls[scenario_id] = transformed
        results.append({
            "name": scenario_name,
            "description": SCENARIO_DESCRIPTIONS.get(scenario_id, ""),
            "status": "generated",
            "path": str(output_path),
        })

    # Generate Excel integration guide
    excel_path = generate_excel_guide(base_xml, scenario_cxmls, req.basePO)

    return GenerateScenariosResponse(
        success=True,
        scenarios=results,
        excelReady=True,
        excelPath=str(excel_path),
    )


SCENARIO_NAMES = {
    "multi-line-split": "Scenario 1: Multi-Line Split",
    "no-supplier-part": "Scenario 2: No Supplier Part Number",
    "adhoc-shipto": "Scenario 3: Ad-hoc Ship-To (ADBUYER)",
}

SCENARIO_DESCRIPTIONS = {
    "multi-line-split": "Single-item PO expanded to multiple line items with split quantities",
    "no-supplier-part": "SupplierPartID removed — buyer-only identification",
    "adhoc-shipto": "Ship-To address replaced with ADBUYER-prefixed ad-hoc address",
}


def apply_transformation(base_xml: str, scenario_id: str) -> str:
    """Apply a transformation scenario to base cXML."""
    parser = etree.XMLParser(recover=True, remove_blank_text=False)
    tree = etree.fromstring(base_xml.encode("utf-8"), parser)

    if scenario_id == "multi-line-split":
        tree = transform_multi_line_split(tree)
    elif scenario_id == "no-supplier-part":
        tree = transform_no_supplier_part(tree)
    elif scenario_id == "adhoc-shipto":
        tree = transform_adhoc_shipto(tree)

    return etree.tostring(tree, pretty_print=True, xml_declaration=True, encoding="UTF-8").decode("utf-8")


def transform_multi_line_split(tree):
    """Split: keep only first ItemOut but create 3 sub-lines from it."""
    items = tree.findall(".//{*}ItemOut") or tree.findall(".//ItemOut")
    if not items:
        return tree
    # Keep first item as template, remove others
    parent = items[0].getparent()
    first_item = deepcopy(items[0])
    for item in items:
        parent.remove(item)

    # Create 3 split lines from original quantity
    orig_qty = int(first_item.get("quantity", "25"))
    quantities = [orig_qty // 3, orig_qty // 3, orig_qty - 2 * (orig_qty // 3)]

    for i, qty in enumerate(quantities, 1):
        new_item = deepcopy(first_item)
        new_item.set("lineNumber", str(i))
        new_item.set("quantity", str(qty))
        desc_el = new_item.find(".//{*}Description") or new_item.find(".//Description")
        if desc_el is not None:
            desc_el.text = f"{(desc_el.text or '').strip()} (Split {i} of 3)"
        parent.append(new_item)

    return tree


def transform_no_supplier_part(tree):
    """Remove SupplierPartID from all items, keep only BuyerPartID."""
    items = tree.findall(".//{*}ItemOut") or tree.findall(".//ItemOut")
    for item in items:
        supplier_parts = item.findall(".//{*}SupplierPartID") or item.findall(".//SupplierPartID")
        for sp in supplier_parts:
            sp.getparent().remove(sp)
        # Ensure BuyerPartID exists
        item_id = item.find(".//{*}ItemID") or item.find(".//ItemID")
        if item_id is not None:
            buyer_part = item_id.find("{*}BuyerPartID") or item_id.find("BuyerPartID")
            if buyer_part is None:
                bp = etree.SubElement(item_id, "BuyerPartID")
                bp.text = f"BUYER-ITEM-{item.get('lineNumber', '0')}"
    return tree


def transform_adhoc_shipto(tree):
    """Replace Ship-To with an ad-hoc ADBUYER-prefixed address."""
    ship_to = tree.find(".//{*}ShipTo") or tree.find(".//ShipTo")
    if ship_to is None:
        return tree
    addr = ship_to.find(".//{*}Address") or ship_to.find(".//Address")
    if addr is not None:
        addr.set("addressID", "ADBUYER-TEMP-001")
    name_el = ship_to.find(".//{*}Name") or ship_to.find(".//Name")
    if name_el is not None:
        name_el.text = "ADBUYER - Temporary Project Site Alpha"
    street = ship_to.find(".//{*}Street") or ship_to.find(".//Street")
    if street is not None:
        street.text = "1200 Construction Blvd, Bldg C"
    city = ship_to.find(".//{*}City") or ship_to.find(".//City")
    if city is not None:
        city.text = "Austin"
    state = ship_to.find(".//{*}State") or ship_to.find(".//State")
    if state is not None:
        state.text = "TX"
    postal = ship_to.find(".//{*}PostalCode") or ship_to.find(".//PostalCode")
    if postal is not None:
        postal.text = "73301"
    return tree


def generate_excel_guide(base_xml: str, scenario_cxmls: dict, po_number: str) -> Path:
    """Generate Excel Integration Guide with scenarios in tabs."""
    wb = openpyxl.Workbook()

    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Overview sheet
    ws = wb.active
    ws.title = "Overview"
    ws.append(["AIE Integration Guide", "", "", ""])
    ws.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M")])
    ws.append(["Base PO", po_number])
    ws.append(["Supplier ANID", "AN01015464739-T"])
    ws.append(["Buyer ANID", "AN01395188653-T"])
    ws.append([])
    ws.append(["Scenario", "Description", "Status", "Doc Type"])
    for sid, name in SCENARIO_NAMES.items():
        ws.append([name, SCENARIO_DESCRIPTIONS.get(sid, ""), "Generated", "OrderRequest"])

    # Per-scenario sheets
    for scenario_id, cxml_content in scenario_cxmls.items():
        sheet_name = scenario_id[:31]  # Excel sheet name limit
        ws_sc = wb.create_sheet(title=sheet_name)

        # Parse and extract fields
        parser = etree.XMLParser(recover=True)
        tree = etree.fromstring(cxml_content.encode("utf-8"), parser)

        # Header row
        headers = ["Field Path", "cXML Element", "Value", "EDI Segment", "EDI Element", "Notes"]
        for col, h in enumerate(headers, 1):
            cell = ws_sc.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Extract key fields
        row = 2
        fields = extract_cxml_fields(tree)
        for field in fields:
            for col, val in enumerate(field, 1):
                cell = ws_sc.cell(row=row, column=col, value=val)
                cell.border = border
            row += 1

        # Auto-width
        for col in ws_sc.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws_sc.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    excel_path = DATA_DIR / f"AIE-Integration-Guide-{po_number}.xlsx"
    wb.save(str(excel_path))
    return excel_path


def extract_cxml_fields(tree) -> List[List[str]]:
    """Extract key cXML fields into tabular format for Excel."""
    fields = []
    # Header fields
    hdr = tree.find(".//{*}OrderRequestHeader") or tree.find(".//OrderRequestHeader")
    if hdr is not None:
        fields.append(["Header/OrderID", "OrderRequestHeader@orderID", hdr.get("orderID", ""), "BEG", "BEG03", "PO Number"])
        fields.append(["Header/OrderDate", "OrderRequestHeader@orderDate", hdr.get("orderDate", ""), "BEG", "BEG05", "PO Date"])
        fields.append(["Header/Type", "OrderRequestHeader@type", hdr.get("type", ""), "BEG", "BEG01", "Purpose Code"])

    total = tree.find(".//{*}Total/{*}Money") or tree.find(".//Total/Money")
    if total is not None:
        fields.append(["Header/Total", "Total/Money", total.text or "", "CTT/AMT", "AMT02", "Total Amount"])
        fields.append(["Header/Currency", "Total/Money@currency", total.get("currency", ""), "CUR", "CUR02", "Currency Code"])

    # ShipTo
    ship_name = tree.find(".//{*}ShipTo//{*}Name") or tree.find(".//ShipTo//Name")
    if ship_name is not None:
        fields.append(["ShipTo/Name", "ShipTo/Address/Name", ship_name.text or "", "N1", "N102", "Ship-To Name"])

    ship_addr = tree.find(".//{*}ShipTo//{*}Address") or tree.find(".//ShipTo//Address")
    if ship_addr is not None:
        fields.append(["ShipTo/AddressID", "ShipTo/Address@addressID", ship_addr.get("addressID", ""), "N1", "N104", "Address ID"])

    # BillTo
    bill_name = tree.find(".//{*}BillTo//{*}Name") or tree.find(".//BillTo//Name")
    if bill_name is not None:
        fields.append(["BillTo/Name", "BillTo/Address/Name", bill_name.text or "", "N1", "N102", "Bill-To Name"])

    # Payment
    pay = tree.find(".//{*}PaymentTerm") or tree.find(".//PaymentTerm")
    if pay is not None:
        fields.append(["Payment/Days", "PaymentTerm@payInNumberOfDays", pay.get("payInNumberOfDays", ""), "ITD", "ITD05", "Net Days"])

    # Line items
    items = tree.findall(".//{*}ItemOut") or tree.findall(".//ItemOut")
    for item in items:
        ln = item.get("lineNumber", "?")
        qty = item.get("quantity", "?")
        fields.append([f"Item[{ln}]/Quantity", f"ItemOut@quantity", qty, "PO1", "PO102", f"Line {ln} qty"])

        sp = item.find(".//{*}SupplierPartID") or item.find(".//SupplierPartID")
        if sp is not None:
            fields.append([f"Item[{ln}]/SupplierPartID", "SupplierPartID", sp.text or "", "PO1", "PO107", "Supplier Part"])

        bp = item.find(".//{*}BuyerPartID") or item.find(".//BuyerPartID")
        if bp is not None:
            fields.append([f"Item[{ln}]/BuyerPartID", "BuyerPartID", bp.text or "", "PO1", "PO109", "Buyer Part"])

        price = item.find(".//{*}UnitPrice/{*}Money") or item.find(".//UnitPrice/Money")
        if price is not None:
            fields.append([f"Item[{ln}]/UnitPrice", "UnitPrice/Money", price.text or "", "PO1", "PO104", "Unit Price"])

        desc = item.find(".//{*}Description") or item.find(".//Description")
        if desc is not None:
            fields.append([f"Item[{ln}]/Description", "Description", (desc.text or "").strip(), "PID", "PID05", "Description"])

        uom = item.find(".//{*}UnitOfMeasure") or item.find(".//UnitOfMeasure")
        if uom is not None:
            fields.append([f"Item[{ln}]/UOM", "UnitOfMeasure", uom.text or "", "PO1", "PO103", "Unit of Measure"])

    return fields


@app.get("/api/aie/po/{po_number}/cxml")
def get_po_cxml(po_number: str):
    """Return stored cXML for a PO."""
    path = DATA_DIR / f"PO-{po_number}.cxml"
    if path.exists():
        return {"success": True, "cxml": path.read_text(encoding="utf-8"), "source": "file"}
    return {"success": False, "error": f"PO {po_number} not found in {DATA_DIR}"}


@app.get("/api/aie/pos")
def list_pos():
    """List all stored PO cXML files."""
    files = list(DATA_DIR.glob("PO-*.cxml"))
    pos = []
    for f in files:
        po_num = f.stem.replace("PO-", "")
        pos.append({"poNumber": po_num, "path": str(f), "size": f.stat().st_size})
    return {"success": True, "pos": pos}


@app.get("/api/aie/retrieve")
def retrieve_from_portal():
    """Live retrieval: scrape POs from supplier portal via Chrome DevTools Protocol."""
    pos = retrieve_pos_sync()
    if pos:
        return {"success": True, "pos": pos, "source": "portal-scrape", "count": len(pos)}
    # Fallback to stored files
    files = list(DATA_DIR.glob("PO-*.cxml"))
    stored = []
    for f in files:
        po_num = f.stem.replace("PO-", "")
        stored.append({"poNumber": po_num, "format": "cXML", "source": "local-file", "status": "Stored"})
    return {"success": True, "pos": stored, "source": "local-fallback", "count": len(stored)}


@app.get("/api/aie/download-excel/{po_number}")
def download_excel(po_number: str):
    """Download the generated Excel integration guide."""
    path = DATA_DIR / f"AIE-Integration-Guide-{po_number}.xlsx"
    if not path.exists():
        # Generate it on the fly
        generate_scenarios(GenerateScenariosRequest(
            basePO=po_number,
            scenarios=["multi-line-split", "no-supplier-part", "adhoc-shipto"]
        ))
    if path.exists():
        return FileResponse(
            path=str(path),
            filename=f"AIE-Integration-Guide-{po_number}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    return {"success": False, "error": "Excel generation failed"}


@app.get("/api/aie/download-cxml/{scenario_id}")
def download_scenario_cxml(scenario_id: str):
    """Download a generated scenario cXML file."""
    path = DATA_DIR / f"scenario-{scenario_id}.cxml"
    if path.exists():
        return FileResponse(path=str(path), filename=f"scenario-{scenario_id}.cxml", media_type="application/xml")
    return {"success": False, "error": f"Scenario {scenario_id} not found"}


class SendTestRequest(BaseModel):
    poNumber: str
    scenarioId: str
    cxmlContent: Optional[str] = None


@app.post("/api/aie/send-test")
def send_test_to_sbn(req: SendTestRequest):
    """Send a test document to SAP Business Network via cXML push."""
    import urllib.request

    # Load scenario cXML or base PO
    cxml_path = DATA_DIR / f"scenario-{req.scenarioId}.cxml"
    if not cxml_path.exists():
        cxml_path = DATA_DIR / f"PO-{req.poNumber}.cxml"
    if not cxml_path.exists():
        return {"success": False, "error": "No cXML file found for this scenario"}

    cxml_content = cxml_path.read_text(encoding="utf-8")

    # Send to Ariba Network cXML endpoint
    url = "https://service.ariba.com/service/transaction/cxml.asp"
    headers = {"Content-Type": "text/xml"}

    try:
        request = urllib.request.Request(url, data=cxml_content.encode("utf-8"), headers=headers)
        with urllib.request.urlopen(request, timeout=30) as response:
            resp_body = response.read().decode("utf-8")
            # Parse response for status
            if "200" in resp_body and "OK" in resp_body:
                return {
                    "success": True,
                    "status": "sent",
                    "response": resp_body[:500],
                    "poNumber": req.poNumber,
                    "scenarioId": req.scenarioId,
                }
            else:
                return {
                    "success": True,
                    "status": "sent-with-warning",
                    "response": resp_body[:500],
                    "poNumber": req.poNumber,
                    "scenarioId": req.scenarioId,
                }
    except Exception as e:
        return {"success": False, "error": str(e), "poNumber": req.poNumber}


@app.get("/api/aie/scenarios")
def list_scenarios():
    """List all generated scenario files."""
    files = list(DATA_DIR.glob("scenario-*.cxml"))
    scenarios = []
    for f in files:
        scenario_id = f.stem.replace("scenario-", "")
        content = f.read_text(encoding="utf-8")
        # Count items
        item_count = content.count("<ItemOut")
        scenarios.append({
            "id": scenario_id,
            "name": SCENARIO_NAMES.get(scenario_id, scenario_id),
            "description": SCENARIO_DESCRIPTIONS.get(scenario_id, ""),
            "items": item_count,
            "size": f.stat().st_size,
            "path": str(f),
        })
    return {"success": True, "scenarios": scenarios}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8100)
